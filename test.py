import openpyxl
import time
import os
import numpy as np
import itertools

import argparse

# name  = 'log_6_2023-5-17-15-43-36_actuator_armed_0.csv'

# lst = name.split(".")
# print(lst)
# llst = lst[0].split("_")
# print(llst)
# llst = llst[3:]
# print(llst)
# new_name = ''
# for i in range(len(llst)):
#     new_name = new_name + '_' + llst[i]

# print(new_name)

# path = 'H:\\SampleData\\HIL\\acce\\TestCase_1_2400000000\\TrueData\\TrueState_data.xlsx'
# wb = openpyxl.load_workbook(path)
# sheet = wb.worksheets[0]
# header = []
# for row in sheet[1]:
#     header.append(row.value)

# print(header)

# a = np.array([[1,2,3],
#               [4,5,6]])

# a = '_slash_mavros_slash_offb_flight_mode_slash_send_data'
# index = a.find('_slash_mavros_slash_')
# print(index)

# lst = list(itertools.repeat('str', 6))
# print(lst)

def get_parse():
    # parse parameters
    parser = argparse.ArgumentParser(description='Dataset process tools')
    parser.add_argument('--restore_path', type=str, default='./ProcessData',
                        help='process data restore path')
    # sub-dataset:
    # default = -1, trans data in the SIL, HIL and Real folder
    # 1 for SIL, 2 for HIL, and 3 for Real
    parser.add_argument('--sub_dataset', type=int, default=-1,
                        help='select the sub_dataset you want')
    # fault type:
    # default = -1, trans all the fault type in the dataset
    # others occasions, please see readme.md
    parser.add_argument('--fault_type', type=int, default=-1,
                        help='select the fault type you want')
    # flight status:
    # default = -1, trans all the flight type in the dataset
    # other occasions, please see readme.md
    parser.add_argument('--flight_status', type=int, default=-1,
                        help='select the flight status you need')
    # trans num:
    # default = -1, trans all the flight cases in the dataset
    # if input other numbers, the program will change the transferred files.
    parser.add_argument('--trans_num', type=int, default=-1,
                        help='the number of cases to transfer')
    return parser

if __name__ == "__main__":
    # parser = get_parse() 

    # args = parser.parse_args()
 
    # print(args.flight_status, args.fault_type)
    START_TIME = time.time()
    path = './SampleData/SIL/hover/accelerometer/TestCase_4_1002000003/TrueData/UAVState_data.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb.worksheets[0]
    data = []
    print(time.time() - START_TIME)
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)
    # 关闭 Excel 文件
    print(time.time() - START_TIME)
    wb.close()

