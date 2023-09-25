# -*- coding: utf-8 -*-
# Copyright (C) rflylab from School of Automation Science and Electrical Engineering, Beihang University.
# All Rights Reserved.
import numpy as np
import csv
import openpyxl
import pandas as pd
import re
import os
import decimal
import itertools
from scipy.interpolate import interp1d

import matplotlib.pyplot as plt
# from windextract import WindExtractor

class CSVFile_extractor:
    def __init__(self, input_path):
        self.project_path = os.path.dirname(__file__)
        self.data_file_path = input_path
        self.timetool = None
        self.str2list_flag = 0
        self.changed_index = 0
        self.target_length = 0

    def reset(self):
        self.str2list_flag = 0
        self.changed_index = 0
        self.target_length = 0
        
    def DataFromCSV(self, file_name, labels, labels_f, path):
        target_path = self.data_file_path + '//' + path
        target_file_path = target_path + '//' + file_name
        all_data = []
        with open(target_file_path, mode='r', encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile)
            header_row = next(csvreader)
            all_data.append(labels)
            # print(header_row)
            col_index = self.label_index_mate(labels, header_row)
            # print(col_index)
            for row in csvreader:
                row_data = []
                for j in range(len(col_index)):
                    trans_data = self.transfer_data_type(row[col_index[j]], labels_f[j])
                    row_data.append(trans_data)
                all_data.append(row_data)
            if len(all_data[1]) == 1:
                all_data = [i for j in range(len(all_data)) for i in all_data[j]]
            return all_data

    def DataFromCSV_Panda(self, file_name, labels, path):
        self.reset()
        target_path = self.data_file_path + '//' + path
        target_file_path = target_path + '//' + file_name
        # df = pd.read_csv(target_file_path, dtype={'col1': int, 'col8': int})
        df = pd.read_csv(target_file_path)
        # print(df.columns.to_list())
        # Timelist = df['frame_id'].to_list()
        # print(type(Timelist[2]))
        # print(len(df.values))
        header = df.columns.to_list()
        col_index = self.label_index_mate(labels, header)
        all_data = []
        all_data.append(labels)
        # for row in df.itertuples():
        #     row_data = []
        #     for i in range(len(row)-1):
        #         row_data.append(row[col_index[i]])
        #     all_data.append(row_data)
        values = df.values
        # all_data += [[values[i][col_i] for col_i in col_index] for i in range(values.shape[0])]
        for i in range(values.shape[0]):
            row_data = []
            for j in range(len(col_index)):
                if type(values[i][col_index[j]]) == str:
                    new_list = self.list_str_check(values[i][col_index[j]])
                    row_data.extend(new_list)
                    # If i == 0, change the header, if not, won't change
                    if i == 0 and self.str2list_flag:
                        self.changed_index = j
                        all_data = self.data_header_adjust(all_data)
                else:
                    row_data.append(values[i][col_index[j]])
            all_data.append(row_data)
        if len(all_data[1]) == 1:
            all_data = [i for j in range(len(all_data)) for i in all_data[j]]
        return all_data

    def DataFromXLSX(self, file_name, labels, path):
        target_path = self.data_file_path + '//' + path
        target_file_path = target_path + '//' + file_name
        all_data = []
        wb = openpyxl.load_workbook(target_file_path, read_only=True)
        sheet = wb.worksheets[0]
        datas = list(sheet.iter_rows(values_only=True))
        # Get header
        header = datas[0]
        # for item in sheet[1]:
        #     header.append(item.value)
        all_data.append(labels)
        col_index = self.label_index_mate(labels, header)
        # Get contents
        data = datas[1:]
        for row in data:
            selecet_row_data = []
            for j in range(len(col_index)):
                selecet_row_data.append(row[col_index[j]])
            all_data.append(selecet_row_data)
        # for row in sheet.iter_rows(min_row=2):
        #     row_data = []
        #     selecet_row_data = []
        #     for cell in row:
        #         row_data.append(cell.value)
        #     for j in range(len(col_index)):
        #         selecet_row_data.append(row_data[col_index[j]])
        #     all_data.append(selecet_row_data)
        wb.close()
        # max_row_num = sheet.max_row
        # for i in range(2, max_row_num+1):
        #     row_data = []
        #     for j in range(len(col_index)):
        #         row_data.append(sheet.cell(row=i, column=1+col_index[j]).value)
        #     all_data.append(row_data)
        if len(all_data[1]) == 1:
            all_data = [i for j in range(len(all_data)) for i in all_data[j]]
        return all_data

    def DataFromXLSX_Panda(self, file_name, labels, path):
        target_path = self.data_file_path + '//' + path
        target_file_path = target_path + '//' + file_name
        all_data = []
        df = pd.read_excel(target_file_path)
        header = df.columns.to_list()
        col_index = self.label_index_mate(labels, header)
        all_data = []
        all_data.append(labels)
        values = df.values
        for i in range(values.shape[0]):
            row_data = []
            for j in range(len(col_index)):
                row_data.append(values[i][col_index[j]])
            all_data.append(row_data)
        if len(all_data[1]) == 1:
            all_data = [i for j in range(len(all_data)) for i in all_data[j]]
        return all_data

    def ConvertXLSXtoCSV(self, file_name, path):
        target_path = self.data_file_path + '//' + path
        file_list = os.listdir(target_path)
        if file_name[:-5] + '.csv' in file_list:
            pass
        else:
            target_file_path = target_path + '//' + file_name
            df = pd.read_excel(target_file_path)
            new_path = target_path + '//' + file_name[:-5] + '.csv'
            df.to_csv(new_path, encoding='utf-8', index=False) 
    
    def label_index_mate(self, labels, header):
        col_index = []
        for i in range(len(labels)):
            index = header.index(labels[i])
            col_index.append(index)
        return col_index

    def get_start_end_time(self, arm_data):
        data_len = len(arm_data)
        first_time = arm_data[1][0]
        last_time = arm_data[data_len-1][0]
        for i in range(1, 10):
            if (arm_data[i][1] != arm_data[i+1][1]) and arm_data[i][1] == '0':
                first_time = arm_data[i+1][0]
                print("New start time!!!")
        for i in range(data_len-1, data_len-11, -1):
            if (arm_data[i][1] != arm_data[i-1][1]) and arm_data[i][1] == '0':
                last_time = arm_data[i-1][0]
                print("New end time!!!")
        return [int(first_time), int(last_time)]

    def transfer_data_type(self, data, format):
        """
        This function is used to trans data type of each labels 
        in different files.
        And is invoked by self.DataFromCSV.
        Now, self.DataFromCSV_Panda is more often use.
        """
        if format == 'str':
            data = str(data)
        elif format == 'int':
            data = int(data)
        elif format == 'float':
            data = float(data)
        elif format == 'listf':
            data = re.findall(r'-?\d+\.?[0-9]*', data)
            data = [float(i) for i in data]
        elif format == 'listi':
            data = re.findall(r'-?\d+\.?[0-9]*', data)
            data = [int(i) for i in data]
        else:  # types do not need to be changed, can be expressed as 'stay' in format!
            pass
        return data

    def list_str_check(self, data):
        if data.find('[') != -1 and data.find(']') != -1:
            new_data = eval(data)
            self.str2list_flag = 1
            self.target_length = len(new_data)
        else:
            new_data = [data]
        return new_data

    def data_header_adjust(self, all_data):
        self.str2list_flag = 0
        name = all_data[0][self.changed_index]
        adjust_name = []
        for i in range(self.target_length):
            adjust_name.append(name + '[' + str(i) + ']')
        all_data[0] = all_data[0][0:self.changed_index] + adjust_name + all_data[0][self.changed_index+1:]
        return all_data

    def generate_data_header(self, label1, label2, info):
        all_file_name = label1 + label2
        header = []
        for i in range(len(all_file_name)):
            for j in range(1, len(info[i])):
                header.append(all_file_name[i] + '_' + info[i][j])
        return header

    def generate_CSV_data(self, DataArray, new_file_name):
        # new_file_path = self.project_path + '//' + new_file_name
        new_file_path = new_file_name
        with open(new_file_path, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(DataArray)

    def folder_finder(self):
        dir_list = os.listdir(self.data_file_path)
        # print(dir_list)
        PX4_file_folder = [f for f in dir_list if f[:4] == "log_"]
        ROS_file_folder = [f for f in dir_list if f[:10] == "rfly_real_"]
        return PX4_file_folder[0], ROS_file_folder[0]

    def folder_finder_SH(self):
        dir_list = os.listdir(self.data_file_path)
        PX4_file_folder = [f for f in dir_list if f[:3] == "Log"]
        GTD_file_folder = [f for f in dir_list if f[:4] == "True"]
        return PX4_file_folder[0], GTD_file_folder[0]

    def judge_ulog_trans(self, path, mode=1):
        # Transfer ulog file into CSVs
        listOfCSVFiles = [f for f in os.listdir(path) if f[-4:] == ".csv"]  # get list of only CSV files in current dir.
        # mode=1, Do not re-generate csv files if there are enough.
        if mode == 1:
            if len(listOfCSVFiles) <= 10:
                os.chdir(path)
                for i in range(len(listOfCSVFiles)):
                    os.remove(listOfCSVFiles[i])
                os.system("for %i in (*); do ulog2csv %i")
                os.chdir(self.project_path)
                print('Transformation Finished !!!')
            else:
                pass
        elif mode == 2:  # Re-generate anyway
            os.chdir(path)
            for i in range(len(listOfCSVFiles)):
                os.remove(listOfCSVFiles[i])
            os.system("for %i in (*); do ulog2csv %i")
            os.chdir(self.project_path)
            print('Transformation Finished !!!')
        # get ulog file name
        PX4_ulg_name = [f for f in os.listdir(path) if f[-4:] == ".ulg"]
        return PX4_ulg_name[0][:-4]

    def RFLY_info_adjust(self, data):
        col_length = len(data[0])
        col_index_list = []
        for i in range(col_length):
            if data[0][i][:9] == 'rfly_ctrl' or data[0][i][:10] == '_rfly_ctrl':
                col_index_list.append(i)
        last_default_index = -1
        last_fault_index = -1
        for i in range(1, len(data)):
            check_result = self.RFLY_fault_id_check(data[i][col_index_list[0]])
            if check_result == "PASS_default":
                last_default_index = i
                continue
            elif check_result == "PASS_fault":
                last_fault_index = i
                continue
            elif check_result == "NeedToChange":
                check_side = self.RFLY_info_side_check(data[i][col_index_list[0]])
                if check_side == 1:
                    if last_fault_index == -1:
                        row_num = self.RFLY_find_next_data(i, data, col_index_list)
                    else:
                        row_num = last_fault_index
                elif check_side == 2:
                    row_num = last_default_index
                for j in range(len(col_index_list)):
                    data[i][col_index_list[j]] = data[row_num][col_index_list[j]]
        return data

    def RFLY_fault_id_check(self, id_str):
        id = eval(id_str)
        if 123450 <= id and id <= 123459 or 123540 <= id and id <= 123549:
            # Check the fault type
            return 'PASS_fault'
        elif id == 1500 or id == 0:
            # This means at the time, there is no fault.
            return 'PASS_default'
        else:
            # The data is regenerated by Function:freq_adjustment, and needed to correct.
            return "NeedToChange"

    def RFLY_info_side_check(self, num):
        num = eval(num)
        num = int(float(num))
        Min_num = 750  # average num of 1500 in real and 0 in simulation
        Max_num = 123549  # maximum fault id in dataset.
        if abs(num - Min_num) > abs(num - Max_num):
            return 1
        else:
            return 2

    def RFLY_find_next_data(self, now_row, data, col_index_list):
        for i in range(now_row, len(data)):
            check_result = self.RFLY_fault_id_check(data[i][col_index_list[0]])
            if check_result == "PASS_fault":
                return i
        return len(data) - 1

    def Ros_file_name_simplify(self, file_name):
        file_num = len(file_name)
        for i in range(file_num):
            if file_name[i].find('_slash_mavros_slash_') != -1:
                file_name[i] = file_name[i][20:]
            elif file_name[i].find('_slash_mavlink_slash_') != -1:
                file_name[i] = file_name[i][21:]
            elif file_name[i].find('_slash_') != -1:
                file_name[i] = file_name[i][7:]
        return file_name

    def find_fly_land_time(self, land_data):
        value = land_data[1][1]
        len_data = len(land_data)
        for i in range(1, len_data):
            if value == land_data[i][1]:
                continue
            else:
                flytime = land_data[i][0]
                break
        fly_index = i
        value = land_data[fly_index+1][2]
        for j in range(fly_index + 1, len_data):
            if value == land_data[j][2]:
                continue
            else:
                landtime = land_data[j][0]
                break
        return int(flytime), int(landtime)

    def cut_data(self, data, begin_time, end_time):
        # First, copy the title
        new_data = [data[0]]
        # Second, get the data within begin and end time
        data_len = len(data)
        for i in range(1, data_len):
            if data[i][0] > begin_time and data[i][0] < end_time:
                new_data.append(data[i])
        return new_data

    def CalcuFreq(self, data):
        data_num = len(data)
        begin_time = data[1][0]
        end_time = data[-1][0]
        constant_time = (end_time - begin_time) / 1e6
        calFreq = data_num / constant_time
        return calFreq

    def adjust_data_length(self, data, target_timevec):
        time_original = data[1:, 0].astype(float)
        expectation_end_time = target_timevec[-1]
        time_interval = target_timevec[-1] - target_timevec[-2]
        now_end_time = time_original[-1]
        repeat_unit = data[-1]  # format: numpy
        i = 0
        while True:
            i = i + 1
            new_end_time = now_end_time + time_interval * i
            repeat_unit[0] = new_end_time
            data = np.concatenate((data, [repeat_unit]), axis=0)
            if new_end_time > expectation_end_time:
                break
        return data

    def freq_adjustment(self, data, target_timevec):
        time_original = data[1:, 0].astype(float)
        # In some cases, the data may not satisfied the length in timesync_status.csv
        # so, we have to repeat the last data a few times to compensate
        if time_original[-1] > target_timevec[-1]:
            pass
        else:
            data = self.adjust_data_length(data, target_timevec)
        time_original = data[1:, 0].astype(float)
        len_col = len(data[0])
        stack_flag = 0
        if len_col != 1:
            stack_flag = 1
        for i in range(1, len_col):
            data_col = data[1:, i]
            type_list_data_col = data_col.tolist()
            if type(type_list_data_col[0]) == str:
                new_data_col = self.str_freq_adjustment(data_col, len(target_timevec))
            else:
                # adjust the frequency of the data
                data_col = data_col.astype(float)
                interpolation_function = interp1d(time_original, data_col)
                new_data_col = interpolation_function(target_timevec)
            if i == 1:
                new_data = new_data_col
                if stack_flag:
                    new_data = np.array([new_data]).T
            else:
                new_data = np.concatenate((new_data, np.array([new_data_col]).T), axis=1)
        return new_data

    def str_freq_adjustment(self, data, target_len):
        now_len = len(data)
        seq_num_list, seq_list, special_label_list = [], [], []
        last_str_seq = data[0]
        seq_num = 1
        # Get the each str and their number.
        for i in range(0, now_len):
            if data[i] == last_str_seq:
                seq_num = seq_num + 1
            else:
                seq_list.append(last_str_seq)
                last_str_seq = data[i]
                seq_num_list.append(seq_num)
                if seq_num == 1:
                    special_label_list.append(1)
                else:
                    special_label_list.append(0)
                seq_num = 1
        seq_list.append(last_str_seq)
        seq_num_list.append(seq_num)
        if seq_num == 1:
            special_label_list.append(1)
        else:
            special_label_list.append(0)
        # print(seq_num_list, seq_list)
        # Calculate each str's ratio in whole list, and their 
        # numbers in new list under the same ratio.
        ratio_list, new_seq_num, new_seq_decimal_num = [], [], []
        for i in range(len(seq_list)):
            ratio_list.append(seq_num_list[i]/len(data))
            new_seq_num.append(ratio_list[i] * target_len)
            new_seq_decimal_num.append(new_seq_num[i] % 1)
        # Get the special process str data num, make sure at least ONE exists.
        for i in range(len(special_label_list)):
            if special_label_list[i] == 1:
                if new_seq_num[i] <= 0.5:
                    new_seq_num[i] = 1
                else:
                    special_label_list[i] = 0
        special_process_num = sum(special_label_list)
        # Adjust the whole number in new list
        minimal_decimal = 1
        minimal_decimal_index = -1
        for i in range(special_process_num):
            for j in range(len(seq_list)):
                if new_seq_decimal_num[j] > 0.5 and new_seq_decimal_num[j] < minimal_decimal:
                    minimal_decimal = new_seq_decimal_num[j]
                    minimal_decimal_index = j
            new_seq_num[minimal_decimal_index] = new_seq_num[minimal_decimal_index] - new_seq_decimal_num[minimal_decimal_index]
            new_seq_decimal_num[minimal_decimal_index] = 0
        # Generate new str data
        new_data = []
        new_seq_int_num = []
        for i in range(len(new_seq_num)):
            new_seq_int_num.append(int(self.decimal_change(new_seq_num[i]))) 
            new_data.extend(list(itertools.repeat(seq_list[i], new_seq_int_num[i])))
        # avoid too many round half up
        now_new_len = sum(new_seq_int_num)
        bias = now_new_len - target_len
        if bias == 0:
            pass
        else:
            if bias > 0:
                minimal_decimal = 1
                minimal_decimal_index = -1
                for i in range(bias):
                    for j in range(len(seq_list)):
                        if new_seq_decimal_num[j] < minimal_decimal and type(new_seq_decimal_num[j]) != int:
                            minimal_decimal = new_seq_decimal_num[j]
                            minimal_decimal_index = j
                    if new_seq_decimal_num[minimal_decimal_index] < 0.5:
                        new_seq_num[minimal_decimal_index] = new_seq_num[minimal_decimal_index] - new_seq_decimal_num[minimal_decimal_index] - 1
                    else:
                        new_seq_num[minimal_decimal_index] = new_seq_num[minimal_decimal_index] - new_seq_decimal_num[minimal_decimal_index]
                    new_seq_decimal_num[minimal_decimal_index] = int(0)
            elif bias < 0:
                max_decimal = 0
                max_decimal_index = -1
                for i in range(abs(bias)):
                    for j in range(len(seq_list)):
                        if new_seq_decimal_num[j] > max_decimal and type(new_seq_decimal_num[j]) != int:
                            max_decimal = new_seq_decimal_num[j]
                            max_decimal_index = j
                    if new_seq_decimal_num[max_decimal_index] > 0.5:
                        new_seq_num[max_decimal_index] = new_seq_num[max_decimal_index] - new_seq_decimal_num[max_decimal_index] + 2
                    else:
                        new_seq_num[max_decimal_index] = new_seq_num[max_decimal_index] - new_seq_decimal_num[max_decimal_index] + 1
                    new_seq_decimal_num[max_decimal_index] = int(0)
            # Generate new str data
            new_data = []
            new_seq_int_num = []
            for i in range(len(new_seq_num)):
                new_seq_int_num.append(int(self.decimal_change(new_seq_num[i]))) 
                new_data.extend(list(itertools.repeat(seq_list[i], new_seq_int_num[i])))
        return new_data

    def decimal_change(self, num):
        num = str(num)
        decimal.getcontext().rounding = "ROUND_HALF_UP"
        new_num = decimal.Decimal(num).quantize(decimal.Decimal("0"))
        return new_num

    def rosdata_timestamp_trans(self, timetool, data):
        data_len = len(data)
        for i in range(1, data_len):
            timenow = data[i][0]
            time_bias = timetool.CalculateTimebiasROS(timenow)
            px4_time = timetool.ROStransPX4(time_bias)
            data[i][0] = px4_time
        return data

    def PX4data_timestamp_trans(self, timetool, data):
        data_len = len(data)
        new_data = np.zeros((data_len, 1), dtype=np.float64)
        for i in range(data_len):
            timenow = data[i]
            time_bias = timetool.CalculateTimebiasPX4(timenow)
            ros_time = timetool.PX4transROS(time_bias)
            new_data[i] = ros_time
        return new_data

    def plot_data(self, time, data):
        """
        time and data needed to be 1-D array vector.
        """
        fig, ax = plt.subplots()
        ax.plot(time, data)
        ax.set_xlabel('time (s)', fontsize=12)
        ax.set_ylabel('adjust_frequency_data', fontsize=12)
        ax.grid()
        plt.show()

if __name__ == "__main__":
    # Wind correlation
    # *******************************************************************
    # Wind process data is moved to windextract.py
    # *******************************************************************
    # Wind_path = 'F://健康评估//数据集论文//实飞//怀来实飞数据//wind'
    # WEcase = WindExtractor(Wind_path)
    # ulogfile_name = PX4_file_folder + '.ulg'
    # xlspath = WEcase.filefinder(ulogfile_name)
    # if xlspath != 100:
    #     WEcase.file_reader(xlspath)
    #     selectedDayWind = WEcase.wind_data_selector(ulogtime, TimeTool)
    #     print(selectedDayWind)
    #     CFEcase.generate_CSV_data(selectedDayWind, 'wind_data.csv')
    # else:
    #     print("No such wind data satisfied! Please re-check!")
    # *******************************************************************
    DataPath = 'F:\\CODE\\Python\\fault_data_process\\SampleData\\Real\\hover\\56_1'
    CFEcase = CSVFile_extractor(DataPath)